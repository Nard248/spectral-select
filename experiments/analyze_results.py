#!/usr/bin/env python3
"""
Analyze Master Run Results
==========================
1. Creates a comprehensive Excel summary with wavelength selections
2. Organizes experiments into configuration-grouped subfolders
"""

import json
import os
import shutil
from pathlib import Path
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import numpy as np


# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESULTS_DIR = PROJECT_ROOT / "results" / "Lichens_Dataset_1_MasterRun"
EXPERIMENTS_DIR = RESULTS_DIR / "experiments"
OUTPUT_EXCEL = RESULTS_DIR / "comprehensive_analysis.xlsx"
GROUPED_DIR = RESULTS_DIR / "grouped_by_config"


def load_results_csv():
    """Load the main results CSV."""
    df = pd.read_csv(RESULTS_DIR / "results.csv")
    # Skip the BASELINE row for analysis
    df = df[df['config'] != 'BASELINE'].copy()
    return df


def load_wavelengths(experiment_folder: str) -> list:
    """Load wavelengths from an experiment folder."""
    wl_path = EXPERIMENTS_DIR / experiment_folder / "wavelengths.json"
    if wl_path.exists():
        with open(wl_path, 'r') as f:
            return json.load(f)
    return []


def extract_config_key(row) -> str:
    """
    Extract configuration key (everything except n_bands_to_select).
    This groups experiments by their configuration setup.
    """
    # Key components that define a configuration
    dim_method = row.get('dimension_selection_method', 'unknown')
    n_dims = int(row.get('n_important_dimensions', 0))
    perturb_method = row.get('perturbation_method', 'unknown')
    norm_method = row.get('normalization_method', 'unknown')
    mag_variant = row.get('magnitude_variant', 'unknown')

    # Create a readable key
    key = f"{dim_method}_dim{n_dims}_{perturb_method}_{norm_method}_{mag_variant}"
    return key


def create_comprehensive_excel(df: pd.DataFrame):
    """Create a comprehensive Excel with multiple sheets."""
    print("Creating comprehensive Excel...")

    wb = Workbook()

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 1: Summary Statistics
    # ─────────────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Summary statistics
    summary_data = [
        ["Master Run Analysis Summary", ""],
        ["", ""],
        ["Total Experiments", len(df)],
        ["Unique Configurations", df.apply(extract_config_key, axis=1).nunique()],
        ["", ""],
        ["Best Accuracy", f"{df['accuracy'].max():.4f}"],
        ["Best F1 Score", f"{df['f1'].max():.4f}"],
        ["Best Kappa", f"{df['kappa'].max():.4f}"],
        ["", ""],
        ["Parameter Ranges", ""],
        ["n_bands_to_select", f"{df['n_bands_to_select'].min():.0f} - {df['n_bands_to_select'].max():.0f}"],
        ["dimension_selection_method", ", ".join(df['dimension_selection_method'].unique())],
        ["n_important_dimensions", ", ".join(map(str, sorted(df['n_important_dimensions'].unique())))],
        ["perturbation_method", ", ".join(df['perturbation_method'].unique())],
        ["normalization_method", ", ".join(df['normalization_method'].unique())],
        ["magnitude_variant", ", ".join(df['magnitude_variant'].unique())],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 50

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 2: All Results with Wavelengths
    # ─────────────────────────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Results")

    # Prepare data with wavelength info
    results_data = []
    for _, row in df.iterrows():
        wavelengths = load_wavelengths(row['config'])

        # Extract top wavelengths
        top_excitations = [w['excitation'] for w in wavelengths[:10]] if wavelengths else []
        top_emissions = [w['emission'] for w in wavelengths[:10]] if wavelengths else []
        top_combos = [w['combination_name'] for w in wavelengths[:5]] if wavelengths else []

        results_data.append({
            'config': row['config'],
            'n_bands': int(row['n_bands_to_select']),
            'accuracy': row['accuracy'],
            'f1': row['f1'],
            'kappa': row['kappa'],
            'ari': row['ari'],
            'nmi': row['nmi'],
            'dim_method': row['dimension_selection_method'],
            'n_dims': int(row['n_important_dimensions']),
            'perturb_method': row['perturbation_method'],
            'norm_method': row['normalization_method'],
            'mag_variant': row['magnitude_variant'],
            'top_5_wavelengths': "; ".join(top_combos),
            'top_excitations': ", ".join(map(str, top_excitations)),
            'top_emissions': ", ".join(map(str, top_emissions)),
        })

    results_df = pd.DataFrame(results_data)
    results_df = results_df.sort_values(['accuracy'], ascending=False)

    # Write headers
    for col_idx, col_name in enumerate(results_df.columns, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_idx, row_data in enumerate(results_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_all.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in ws_all.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_all.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 3: Top Configurations by Accuracy
    # ─────────────────────────────────────────────────────────────────────
    ws_top = wb.create_sheet("Top 100 Configs")

    top_100 = results_df.head(100)
    for col_idx, col_name in enumerate(top_100.columns, 1):
        cell = ws_top.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row_data in enumerate(top_100.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_top.cell(row=row_idx, column=col_idx, value=value)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 4: Accuracy by n_bands (aggregated)
    # ─────────────────────────────────────────────────────────────────────
    ws_nbands = wb.create_sheet("Accuracy vs n_bands")

    # Aggregate by n_bands
    nbands_stats = df.groupby('n_bands_to_select').agg({
        'accuracy': ['mean', 'std', 'max', 'min'],
        'f1': ['mean', 'max'],
        'kappa': ['mean', 'max']
    }).round(4)
    nbands_stats.columns = ['_'.join(col) for col in nbands_stats.columns]
    nbands_stats = nbands_stats.reset_index()

    for col_idx, col_name in enumerate(nbands_stats.columns, 1):
        cell = ws_nbands.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row_data in enumerate(nbands_stats.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_nbands.cell(row=row_idx, column=col_idx, value=value)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 5: Best Config per n_bands
    # ─────────────────────────────────────────────────────────────────────
    ws_best = wb.create_sheet("Best per n_bands")

    best_per_nbands = df.loc[df.groupby('n_bands_to_select')['accuracy'].idxmax()]
    best_per_nbands = best_per_nbands[['n_bands_to_select', 'accuracy', 'f1', 'kappa',
                                        'dimension_selection_method', 'n_important_dimensions',
                                        'perturbation_method', 'normalization_method',
                                        'magnitude_variant', 'config']]
    best_per_nbands = best_per_nbands.sort_values('n_bands_to_select')

    for col_idx, col_name in enumerate(best_per_nbands.columns, 1):
        cell = ws_best.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row_data in enumerate(best_per_nbands.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_best.cell(row=row_idx, column=col_idx, value=value)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 6: Wavelength Frequency Analysis
    # ─────────────────────────────────────────────────────────────────────
    ws_wl = wb.create_sheet("Wavelength Frequency")

    # Count how often each wavelength combo appears in top selections
    wl_counts = defaultdict(lambda: {'count': 0, 'avg_rank': 0, 'ranks': []})

    for _, row in df.iterrows():
        wavelengths = load_wavelengths(row['config'])
        for wl in wavelengths:
            combo = wl['combination_name']
            wl_counts[combo]['count'] += 1
            wl_counts[combo]['ranks'].append(wl['rank'])
            wl_counts[combo]['excitation'] = wl['excitation']
            wl_counts[combo]['emission'] = wl['emission']

    # Calculate average rank
    wl_data = []
    for combo, data in wl_counts.items():
        wl_data.append({
            'combination': combo,
            'excitation': data['excitation'],
            'emission': data['emission'],
            'selection_count': data['count'],
            'avg_rank': np.mean(data['ranks']),
            'min_rank': min(data['ranks']),
            'max_rank': max(data['ranks'])
        })

    wl_df = pd.DataFrame(wl_data)
    wl_df = wl_df.sort_values('selection_count', ascending=False)

    for col_idx, col_name in enumerate(wl_df.columns, 1):
        cell = ws_wl.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row_data in enumerate(wl_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_wl.cell(row=row_idx, column=col_idx, value=value)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 7: Configuration Comparison
    # ─────────────────────────────────────────────────────────────────────
    ws_config = wb.create_sheet("Config Comparison")

    # Add config key to dataframe
    df['config_key'] = df.apply(extract_config_key, axis=1)

    config_stats = df.groupby('config_key').agg({
        'accuracy': ['mean', 'std', 'max'],
        'f1': ['mean', 'max'],
        'n_bands_to_select': 'count'
    }).round(4)
    config_stats.columns = ['acc_mean', 'acc_std', 'acc_max', 'f1_mean', 'f1_max', 'n_experiments']
    config_stats = config_stats.sort_values('acc_max', ascending=False).reset_index()

    for col_idx, col_name in enumerate(config_stats.columns, 1):
        cell = ws_config.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row_data in enumerate(config_stats.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_config.cell(row=row_idx, column=col_idx, value=value)

    # Save workbook
    wb.save(OUTPUT_EXCEL)
    print(f"✓ Saved comprehensive Excel to: {OUTPUT_EXCEL}")

    return df


def create_grouped_folders(df: pd.DataFrame):
    """Create folders grouped by configuration (everything except n_bands)."""
    print("\nCreating grouped folders...")

    # Add config key if not present
    if 'config_key' not in df.columns:
        df['config_key'] = df.apply(extract_config_key, axis=1)

    # Create base directory
    GROUPED_DIR.mkdir(exist_ok=True)

    # Group experiments
    groups = df.groupby('config_key')

    print(f"Found {len(groups)} unique configurations")

    for config_key, group_df in groups:
        # Create folder for this configuration
        config_folder = GROUPED_DIR / config_key
        config_folder.mkdir(exist_ok=True)

        # Create a summary CSV for this configuration
        summary_data = []

        for _, row in group_df.iterrows():
            exp_name = row['config']
            src_folder = EXPERIMENTS_DIR / exp_name
            dst_folder = config_folder / exp_name

            # Copy experiment folder (create symlink to save space)
            if src_folder.exists() and not dst_folder.exists():
                # Use symlink instead of copy to save disk space
                dst_folder.symlink_to(src_folder)

            # Add to summary
            summary_data.append({
                'n_bands': int(row['n_bands_to_select']),
                'accuracy': row['accuracy'],
                'f1': row['f1'],
                'kappa': row['kappa'],
                'ari': row['ari'],
                'nmi': row['nmi'],
                'folder': exp_name
            })

        # Save configuration summary
        summary_df = pd.DataFrame(summary_data)
        summary_df = summary_df.sort_values('n_bands')
        summary_df.to_csv(config_folder / 'summary.csv', index=False)

        # Also save as Excel with chart data
        summary_df.to_excel(config_folder / 'summary.xlsx', index=False)

    print(f"✓ Created {len(groups)} grouped folders in: {GROUPED_DIR}")

    # Create an index of all configurations
    index_data = []
    for config_key, group_df in groups:
        index_data.append({
            'config_key': config_key,
            'n_experiments': len(group_df),
            'best_accuracy': group_df['accuracy'].max(),
            'best_n_bands': int(group_df.loc[group_df['accuracy'].idxmax(), 'n_bands_to_select']),
            'accuracy_at_10_bands': group_df[group_df['n_bands_to_select'] == 10]['accuracy'].values[0] if 10 in group_df['n_bands_to_select'].values else None,
            'accuracy_at_50_bands': group_df[group_df['n_bands_to_select'] == 50]['accuracy'].values[0] if 50 in group_df['n_bands_to_select'].values else None,
            'accuracy_at_100_bands': group_df[group_df['n_bands_to_select'] == 100]['accuracy'].values[0] if 100 in group_df['n_bands_to_select'].values else None,
        })

    index_df = pd.DataFrame(index_data)
    index_df = index_df.sort_values('best_accuracy', ascending=False)
    index_df.to_csv(GROUPED_DIR / 'configurations_index.csv', index=False)
    index_df.to_excel(GROUPED_DIR / 'configurations_index.xlsx', index=False)

    print(f"✓ Created configuration index")


def print_insights(df: pd.DataFrame):
    """Print key insights from the analysis."""
    print("\n" + "="*70)
    print("KEY INSIGHTS FROM MASTER RUN")
    print("="*70)

    # Add config key
    if 'config_key' not in df.columns:
        df['config_key'] = df.apply(extract_config_key, axis=1)

    # Best overall
    best_row = df.loc[df['accuracy'].idxmax()]
    print(f"\n🏆 BEST OVERALL ACCURACY: {best_row['accuracy']:.4f}")
    print(f"   Configuration: {best_row['config']}")
    print(f"   n_bands: {int(best_row['n_bands_to_select'])}")

    # Best by dimension method
    print("\n📊 BEST BY DIMENSION METHOD:")
    for method in df['dimension_selection_method'].unique():
        subset = df[df['dimension_selection_method'] == method]
        best = subset.loc[subset['accuracy'].idxmax()]
        print(f"   {method}: {best['accuracy']:.4f} (n_bands={int(best['n_bands_to_select'])})")

    # Best by normalization
    print("\n📊 BEST BY NORMALIZATION:")
    for method in df['normalization_method'].unique():
        subset = df[df['normalization_method'] == method]
        best = subset.loc[subset['accuracy'].idxmax()]
        print(f"   {method}: {best['accuracy']:.4f} (n_bands={int(best['n_bands_to_select'])})")

    # Optimal n_bands range
    print("\n📊 ACCURACY BY N_BANDS (top performers):")
    for n in [5, 10, 20, 30, 50, 100, 150, 180]:
        if n in df['n_bands_to_select'].values:
            subset = df[df['n_bands_to_select'] == n]
            print(f"   n={n:3d}: max={subset['accuracy'].max():.4f}, mean={subset['accuracy'].mean():.4f}")

    # Find the "sweet spot" - where accuracy plateaus
    nbands_means = df.groupby('n_bands_to_select')['accuracy'].max().sort_index()
    print(f"\n📈 ACCURACY PROGRESSION (max per n_bands):")
    for n, acc in list(nbands_means.items())[:10]:
        print(f"   n={int(n):3d}: {acc:.4f}")
    print("   ...")
    for n, acc in list(nbands_means.items())[-5:]:
        print(f"   n={int(n):3d}: {acc:.4f}")

    # Top 5 configurations
    print("\n🏆 TOP 5 CONFIGURATIONS:")
    top5 = df.nlargest(5, 'accuracy')
    for i, (_, row) in enumerate(top5.iterrows(), 1):
        print(f"   {i}. Acc={row['accuracy']:.4f} | {row['config']}")


def main():
    print("="*70)
    print("MASTER RUN ANALYSIS")
    print("="*70)

    # Load data
    print("\nLoading results...")
    df = load_results_csv()
    print(f"Loaded {len(df)} experiments")

    # Create comprehensive Excel
    df = create_comprehensive_excel(df)

    # Create grouped folders
    create_grouped_folders(df)

    # Print insights
    print_insights(df)

    print("\n" + "="*70)
    print("ANALYSIS COMPLETE")
    print("="*70)
    print(f"\nOutputs:")
    print(f"  📊 Excel: {OUTPUT_EXCEL}")
    print(f"  📁 Grouped folders: {GROUPED_DIR}")


if __name__ == "__main__":
    main()
