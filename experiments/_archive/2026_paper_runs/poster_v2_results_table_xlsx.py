#!/usr/bin/env python3
"""Generate the merged Lichens + Pepsin results table as a styled XLSX.
Numbers come from the actual run CSVs:
  Lichens : results/Lichens_Dataset_1_MasterRun/results.csv      (baseline 192 bands)
  Pepsin  : results/Collagen_Pepsin_Normalized/results.csv       (baseline 158 bands)

Output: Showcase_Poster/POSTER_v2_results_table.xlsx
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "Showcase_Poster" / "POSTER_v2_results_table.xlsx"

# ----------------------------------------------------------------------
# Real numbers extracted from the master-run CSVs.
# Each row: (Dataset, Task, Total bands, K, Acc. full, Acc. sel., F1 sel., note)
# ----------------------------------------------------------------------
ROWS = [
    # Lichens
    ("Lichens (TPAMI)",      "8-class substrate ID",
     192,  4,  0.8815, 0.8812, 0.8826, "matches full cube with 4 bands"),
    ("Lichens (TPAMI)",      "8-class substrate ID",
     192, 13,  0.8815, 0.9016, 0.9024, "+2.0 pp over full cube"),
    ("Lichens (TPAMI)",      "8-class substrate ID",
     192, 80,  0.8815, 0.9523, 0.9523, "peak accuracy (+7.1 pp)"),
    # Pepsin
    ("Pepsin–Collagen (IASIM 2026)", "treatment detection",
     158,  5,  0.7978, 0.8085, 0.7967, "+1.1 pp with 5 bands"),
    ("Pepsin–Collagen (IASIM 2026)", "treatment detection",
     158, 20,  0.7978, 0.8496, 0.8500, "+5.2 pp"),
    ("Pepsin–Collagen (IASIM 2026)", "treatment detection",
     158, 30,  0.7978, 0.8559, 0.8550, "peak accuracy (+5.8 pp)"),
]

HEADERS = [
    "Dataset", "Task", "Total bands", "K selected", "Compression",
    "Acc. (full cube)", "Acc. (selected)", "Δ acc. (pp)", "F1 (selected)", "Notes",
]

# ----------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------
NAVY = "1A2332"
NAVY_LIGHT = "E8ECF2"
ZEBRA = "F4F6FA"
ACCENT = "3A86C8"
LICHEN_TINT = "EAF1F9"
PEPSIN_TINT = "FBEFEA"

thin = Side(border_style="thin", color="B9C4D0")
medium = Side(border_style="medium", color=NAVY)
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

font_title = Font(name="Calibri", size=16, bold=True, color=NAVY)
font_subtitle = Font(name="Calibri", size=11, italic=True, color="5A6B7D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11, color="1A2332")
font_body_b = Font(name="Calibri", size=11, bold=True, color="1A2332")
font_note = Font(name="Calibri", size=10, italic=True, color="5A6B7D")

fill_header = PatternFill("solid", fgColor=NAVY)
fill_zebra = PatternFill("solid", fgColor=ZEBRA)
fill_lichen = PatternFill("solid", fgColor=LICHEN_TINT)
fill_pepsin = PatternFill("solid", fgColor=PEPSIN_TINT)

centre = Alignment(horizontal="center", vertical="center")
left_a = Alignment(horizontal="left", vertical="center", indent=1)
right_a = Alignment(horizontal="right", vertical="center", indent=1)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # ---- title + subtitle ------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(HEADERS))
    ws.cell(row=1, column=1,
            value="spectral-select · merged results across datasets")
    ws.cell(row=1, column=1).font = font_title
    ws.cell(row=1, column=1).alignment = centre
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=len(HEADERS))
    ws.cell(row=2, column=1,
            value=("Same 3-D CAE + latent-perturbation + MMR pipeline "
                   "applied to two unrelated datasets. "
                   "Baseline = full-cube kNN; selected = kNN on the K bands "
                   "ranked by spectral-select."))
    ws.cell(row=2, column=1).font = font_subtitle
    ws.cell(row=2, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ---- header row ------------------------------------------------------
    header_row = 4
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = centre
        c.border = border_header
    ws.row_dimensions[header_row].height = 24

    # ---- data rows -------------------------------------------------------
    for i, (dataset, task, tot, K, acc_full, acc_sel, f1_sel, note) in enumerate(
            ROWS, start=header_row + 1):
        compression = (tot - K) / tot
        delta_pp = (acc_sel - acc_full) * 100

        is_lichen = dataset.startswith("Lichens")
        tint = fill_lichen if is_lichen else fill_pepsin

        values = [
            dataset, task, tot, K,
            compression, acc_full, acc_sel, delta_pp, f1_sel, note,
        ]
        aligns = [left_a, left_a, centre, centre,
                  centre, centre, centre, centre, centre, left_a]
        for j, (v, al) in enumerate(zip(values, aligns), start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = font_body_b if j in (1, 4, 8) else font_body
            c.alignment = al
            c.fill = tint
            c.border = border_thin

        # number formats
        ws.cell(row=i, column=5).number_format = "0.0%"      # compression
        ws.cell(row=i, column=6).number_format = "0.000"     # acc full
        ws.cell(row=i, column=7).number_format = "0.000"     # acc sel
        ws.cell(row=i, column=8).number_format = "+0.0;-0.0;0.0"  # Δ pp
        ws.cell(row=i, column=9).number_format = "0.000"     # f1
        ws.row_dimensions[i].height = 22

    # ---- footnote --------------------------------------------------------
    note_row = header_row + 1 + len(ROWS) + 1
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row,   end_column=len(HEADERS))
    ws.cell(row=note_row, column=1, value=(
        "Source: results/Lichens_Dataset_1_MasterRun/results.csv (Lichens, baseline = 192 bands), "
        "results/Collagen_Pepsin_Normalized/results.csv (Pepsin, baseline = 158 bands). "
        "kNN classifier; best-of-config per K reported. "
        "Selection: AE-perturbation + MMR (λ = 0.5) with max_per_excitation normalisation."))
    ws.cell(row=note_row, column=1).font = font_note
    ws.cell(row=note_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[note_row].height = 36

    # ---- column widths ---------------------------------------------------
    widths = {
        "A": 32,  # Dataset
        "B": 24,  # Task
        "C": 13,  # Total bands
        "D": 12,  # K selected
        "E": 14,  # Compression
        "F": 18,  # Acc full
        "G": 18,  # Acc selected
        "H": 14,  # Δ pp
        "I": 16,  # F1
        "J": 36,  # Notes
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
